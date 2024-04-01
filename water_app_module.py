# water_app_module.py

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import json
from math import ceil
import requests
from firebase import firebase

pdfmetrics.registerFont(TTFont('ArialUnicode', 'arial.ttf'))


class WaterApp:
    WORK_FILES_FOLDER = '/work_files'
    def __init__(self, parent_frame,initial_width=800, initial_height=600):
        self.parent_frame = parent_frame
        self.entry_vars = []
        self.entry_widgets = []
        self.saved_values = {}
        
        self.scroll_canvas = tk.Canvas(self.parent_frame)
        self.initial_width = initial_width
        self.initial_height = initial_height
        self.button_state = tk.StringVar(value='submit')

        self.create_widgets()

    def load_data(self, file_path, sheet_name='Sheet1'):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook[sheet_name]
            data = [(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value, sheet.cell(
                row=row, column=3).value, sheet.cell(row=row, column=4).value) for row in range(2, sheet.max_row + 1)]
            return data
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"The Excel file '{file_path}' is not found.")
            exit()

    def create_widgets(self):
        data = self.load_data('res/счетчики_вода.xlsx')

        for widget in self.parent_frame.winfo_children():
            widget.destroy()

        for i, (value1, value2, value3, value4) in enumerate(data):
            self.create_widget(i, value1, value2, value3, value4)

    def create_widget(self, row, value1, value2, value3, value4):
        label_text = f"{value1} - {value2} - {value4}" if value1 is not None else ''
        label = ttk.Label(self.parent_frame, text=label_text, font='Arial 11 bold')
        label.grid(row=row, column=0, padx=5, pady=5, sticky='w')

        entry_var = tk.StringVar(value=self.saved_values.get(f"{value1}_{value2}_{value4}", ""))
        self.entry_vars.append(entry_var)

        is_preloaded = entry_var.get() != ""
        entry_bg_color = 'lightgrey' if is_preloaded else 'white'
        entry_font = 'Arial 11 bold' if not is_preloaded else 'Arial 10 italic'

        entry = ttk.Entry(self.parent_frame, textvariable=entry_var, font=entry_font, background=entry_bg_color)
        entry.grid(row=row, column=1, padx=5, pady=5, sticky='w')
        self.entry_widgets.append(entry)

        label_text_comments = f"{value3}"
        label_comments = ttk.Label(self.parent_frame, text=label_text_comments, font='Arial 10 bold')
        label_comments.grid(row=row, column=2, padx=5, pady=5, sticky='w')

        entry.bind('<Button-1>', lambda event, entry=entry, is_preloaded=is_preloaded: self.on_entry_click(event, entry, is_preloaded))

    def on_entry_click(self, event, entry, is_preloaded):
        if is_preloaded:
            entry.config(font='Arial 11 bold', background='white')
            entry.icursor(0)

    def update_gui(self):
        self.entry_vars = []
        self.entry_widgets = []

        data = self.load_data('res/счетчики_вода.xlsx')

        for widget in self.parent_frame.winfo_children():
            widget.destroy()

        for i, (value1, value2, value3, value4) in enumerate(data):
            self.create_widget(i, value1, value2, value3, value4)

    def on_submit(self, submit_button):
        self.generate_pdf('test.pdf')

        self.save_values_to_file('saved_values1.json', {f"{value1}_{value2}_{value4}": entry_var.get()
                                                        for entry_var, (value1, value2, _, value4) in zip(self.entry_vars, self.data)})

        for entry in self.entry_widgets:
            entry.config(state="disabled")

        button_style = ttk.Style()
        button_style.configure('Submit.TButton', background='red', foreground='black')

        if self.button_state.get() == 'submit':
            submit_button.config(text="Выход", command=self.parent_frame.destroy)
            self.button_state.set('exit')
        else:
            submit_button.config(text="Отправить", command=lambda: self.on_submit(submit_button))
            self.button_state.set('submit')

        #--------------FIREBASE
        # Generate PDF
        pdf_filename = 'test.pdf'
        self.generate_pdf(pdf_filename)

        # Upload PDF to Firebase Storage
        firebase_config = {
            'apiKey': 'your-api-key',
            'authDomain': 'your-project-id.firebaseapp.com',
            'storageBucket': 'your-project-id.appspot.com',
        }

        fb = firebase.FirebaseApplication('https://your-firebase-project.firebaseio.com/', None)
        storage_url = fb.get('/your-storage-url', None)  # Replace with your actual storage URL

        storage_url += pdf_filename

        try:
            with open(pdf_filename, 'rb') as file:
                files = {'file': (pdf_filename, file)}
                response = requests.post(storage_url, files=files)
                response.raise_for_status()
                print("PDF uploaded to Firebase Storage successfully")
        except requests.exceptions.RequestException as e:
            print(f"Error uploading PDF to Firebase Storage: {e}")

    def generate_pdf(self, filename):
        current_datetime = datetime.now().strftime("%d%m%Y")
        pdf_filename = f"Показания_водоснабжение_{current_datetime}.pdf"

        pdf_canvas_obj = pdf_canvas.Canvas(pdf_filename)

        # Use Arial Unicode MS font
        pdf_canvas_obj.setFont("ArialUnicode", 16)

        # Header
        pdf_canvas_obj.drawCentredString(
            300, 750, f'Показания водоснабжения за: {datetime.now().strftime("%B %Y")}')

        pdf_canvas_obj.setFont("ArialUnicode", 15)
        pdf_canvas_obj.drawCentredString(
            300, 810, f"Протокол показаний счетчиков водоснабжения. ")
        pdf_canvas_obj.drawCentredString(
            300, 780, f"Адрес: Мос. обл., Солнечногорский р-н., пос. Поварово, мкр. Поваровка-12")

        entries_per_page = 25  # Adjust as needed
        total_entries = len(self.entry_vars)

        for page_number in range(ceil(total_entries / entries_per_page)):
            start_index = page_number * entries_per_page
            end_index = min((page_number + 1) * entries_per_page, total_entries)

            # Content
            y = 700  # Starting y-coordinate
            for i in range(start_index, end_index):
                entry_var, = self.entry_vars[i],
                value1, value2, value3, value4 = self.load_data('res/счетчики_вода.xlsx')[i]

                user_input = entry_var.get() or "Нет показаний"

                pdf_canvas_obj.setFont("ArialUnicode", 12)
                pdf_canvas_obj.drawString(
                    100, y, f"Показания счетчика - {value1} {value2} {value4} - : {user_input}")
                y -= 20  # Adjust the vertical spacing

            # Signature line
            signature_y = 100  # Adjust this value for proper vertical placement
            pdf_canvas_obj.setFont("ArialUnicode", 12)
            pdf_canvas_obj.drawString(100, signature_y - 50,
                                      "__________________________")
            pdf_canvas_obj.drawString(100, signature_y - 70,
                                      "Подпись ООО \"Компания Мебельстайл\"")

            # Add page number to each page
            pdf_canvas_obj.setFont("ArialUnicode", 8)
            pdf_canvas_obj.drawString(
                500, 50, f'Страница {page_number + 1} из {ceil(total_entries / entries_per_page)}')

            # Add a new page if not the last page
            if page_number < (total_entries // entries_per_page):
                pdf_canvas_obj.showPage()

        try:
            pdf_canvas_obj.save()
            print(f"PDF generated: {pdf_filename}")
        except Exception as e:
            print(f"Error during PDF generation: {e}")
       

    def save_values_to_file(self, file_path, values):
        file_path_in_work_files = os.path.join(self.WORK_FILES_FOLDER, file_path)
        with open(file_path_in_work_files, 'w') as file:
            json.dump(values, file)

    def load_values_from_file(self, file_path='saved_values1.json'):
        file_path_in_work_files = os.path.join(self.WORK_FILES_FOLDER, file_path)
        try:
            with open(file_path_in_work_files, 'r') as file:
                return json.load(file)
        except FileNotFoundError:
            return {}

    def clear_default_value(self, event, entry_var, default_value):
        if entry_var.get() == default_value:
            entry_var.set("")

    def on_exit(self):
        self.parent_frame.destroy()

    def on_resize(self, event):
        self.parent_frame.after(500, self.update_after_resize)
        
    def update_after_resize(self):
        # Get the new size of the window
        new_width = self.scroll_canvas.winfo_width()
        new_height = self.scroll_canvas.winfo_height()

        # Calculate the scale factor
        scale_factor = min(new_width / self.initial_width, new_height / self.initial_height)

        # Update the scale factor and redraw the content
        self.scroll_canvas.scale("all", 0, 0, scale_factor, scale_factor)